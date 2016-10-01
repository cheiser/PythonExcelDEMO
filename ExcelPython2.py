import openpyxl
import datetime
from openpyxl.cell import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook("testPython.xlsx") # load the workbook (i.e. excel sheet)
iWs = wb.active # wb2.create_sheet()
# wb = openpyxl.load_workbook("updatedTestPython2.xlsx") # load the workbook (i.e. excel sheet)
wb2 = openpyxl.Workbook()
oWs = wb2.active # wb2.create_sheet()

# type-correct the work sheet
# iWs - the input worksheet
# oWs - the output worksheet
def type_correct(iWs, oWs):
    for row in iWs["A1":str(get_column_letter(iWs.max_column)) + str(iWs.max_row)]:
        for cell in row:
            if(type(cell.value) == unicode):
                try:
                    tempValue = float(cell.value)
                    print "string float", cell.coordinate
                    # cell.set_explicit_value(value=1.3, data_type=cell.TYPE_NUMERIC)
                    # oWs[cell.coordinate].set_explicit_value(value=1.3, data_type=cell.TYPE_NUMERIC)
                    oWs[cell.coordinate] = tempValue # place the corrected float at same position in output
                except Exception as e:
                    print "NOT A FLOAT", cell.coordinate, e
                    oWs[cell.coordinate] = cell.value
            elif(type(cell.value) == datetime.time):
                print "datetime, save as is", cell.coordinate
                oWs[cell.coordinate] = cell.value
            else:
                print "float?", cell.coordinate
                oWs[cell.coordinate] = float(cell.value)


type_correct(iWs, oWs)

print "THE TYPE AT C2 IS ", type(oWs["C2"].value), " - ", oWs["C2"].value
wb2.save("updatedTestPython2.xlsx")
