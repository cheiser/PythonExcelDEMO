import openpyxl
import datetime
from openpyxl.cell import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook("testPython.xlsx") # load the workbook (i.e. excel sheet)
# wb = openpyxl.load_workbook("updatedTestPython2.xlsx") # load the workbook (i.e. excel sheet)
wb2 = openpyxl.Workbook()
ws = wb2.active # wb2.create_sheet()

sheet_names = wb.get_sheet_names() # get the sheet names
specificSheet = wb.get_sheet_by_name(sheet_names[0]) # gets a specific sheet by the specified name
print specificSheet.title # the title of the sheet

activeSheet = wb.active # gets the active sheet

cell = specificSheet["B1"] # cell B 1
print specificSheet["A1"].value # value at column A index 1
print "ROW: ", str(cell.row), " COLUMN: ", str(cell.column), " VALUE: ", str(cell.value)

# we can also access a cell by
otherCell = activeSheet.cell(row=1, column=2) # column 2 = B, row 1 = 1, i.e. B1
print "THE SHEET HAS: ", str(activeSheet.max_row), " ROWS AND ", str(activeSheet.max_column), \
        "COLUMNS"

print "THE COLUMN LETTER AT 1 IS ", get_column_letter(1), " INDEX FOR B IS ", column_index_from_string("B")

sheetTuple = tuple(activeSheet["A1":"C3"]) # gets all cells between A1 to C3
for rowOfCellObjects in activeSheet["A1":"C3"]:
    for cellObj in rowOfCellObjects:
        print cellObj.coordinate, " ", cellObj.value

    print "----------------------------"

sheetColumn1 = activeSheet.columns[1] # for cellObj in activeSheet.columns[1]: cellObj.value

mRow = activeSheet.max_row
mCol = activeSheet.max_column
endRowCol = str(get_column_letter(mCol)) + str(mRow)
print "endRowCol: ", endRowCol
for rowOfCellObjects in activeSheet["A1":endRowCol]:
# for rowOfCellObjects in activeSheet["A1":"F6"]:
    for cellObj in rowOfCellObjects:
        print "SETTING VALUE AT ", cellObj.coordinate, " WITH TYPE: ", type(cellObj.value)
        if(type(cellObj.value) == datetime.time):
            # set default value 1.0
            print "ENTERING DATE VALUE", cellObj.value
            ws[cellObj.coordinate] = 1.0
        elif(type(cellObj.value) == unicode):
            # check if string actually is float then enter as float
            try:
                tempValue = float(cellObj.value)
                print "ACTUALLY FLOAT VALUE!!!", tempValue
                ws[cellObj.coordinate].set_explicit_value(value=tempValue, data_type=cellObj.TYPE_NUMERIC)
                # ws[cellObj.coordinate] = 1.3 # tempValue
                print "value-type: ", ws[cellObj.coordinate].value, "-", type(ws[cellObj.coordinate].value)
                print "AT: ", cellObj.coordinate, ": ", type(ws["C2"].value)
            except:
                pass
            print "ENTERING STRING VALUE", cellObj.value
            ws[cellObj.coordinate] = cellObj.value.encode("utf-8") # cast the value to proper utf-8
        else:
            print "ENTERING FLOAT VALUE", cellObj.value
            ws[cellObj.coordinate] = float(cellObj.value)
        cellObj.value = float(1.0)
        # cellObj.style.number_format = "????????.??????"
        if type(cellObj.value) != type(1.0):
            print "forcing float"
            cellObj.set_explicit_value(value=1.0, data_type=cellObj.TYPE_NUMERIC)
        print "after assignment: ", type(cellObj.value)


print "THE TYPE AT C2 IS ", type(ws["C2"].value)
# wb.save("updatedTestPython.xlsx")

wb2.save("updatedTestPython2.xlsx")
